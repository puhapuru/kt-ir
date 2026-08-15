#!/usr/bin/env python3
"""IR Factsheet 에서 가입자 수와 ARPU 를 뽑는다.

    .venv/bin/python scripts/extract_factsheet.py --report
    .venv/bin/python scripts/extract_factsheet.py

**형식이 세 세대다.** 25년치라 그렇다. 지금 다루는 것은 **현대 형식**
(`Fixed(M)` + `Wireless(Q)` 두 시트)이고, 나머지 둘은 아직 안 건드렸다 —
`--report` 가 몇 건이 그렇게 빠졌는지 알려 준다.

    현대   Fixed(M) + Wireless(Q)        영문 표기, 값은 명 단위
    중기   Factsheet(Kor)/(Eng)          한글 표기, 값이 **천명**
    초기   단일 시트 `KT Factsheet`       영문, 항목 구성이 다름

## 시트 두 장이 주기가 다르다 ★

    Wireless(Q)   **분기**  — 가입자·MNO·MVNO·5G·해지율·ARPU
    Fixed(M)      **월**    — 유선전화·초고속인터넷·IPTV

한 파일 안에서 주기가 갈리므로 시트마다 따로 읽는다.

## 여기도 머리글이 거짓이다

`Subscribers ('000)` 라고 적혀 있는데 값은 `20,623,074` — **명 단위**다.
Financial Indicator 의 `(Billion KRW)` 와 같은 함정이라, 여기서도 값 크기로
판정한다(KT 무선 가입자는 1,500만~3,500만 명 선).

## 날짜 칸이 문자열이 아니다

월 시트의 머리글은 엑셀 **날짜 값**으로 온다(`2018-01-01 00:00`). 옛 `.xls`
에서는 **일련번호**(`38355.0`)로 온다. 둘 다 받는다.
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
import warnings
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import factstore as FS  # noqa: E402

warnings.filterwarnings("ignore")

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl 이 없다:  .venv/bin/pip install openpyxl xlrd")
try:
    import xlrd
except ImportError:
    xlrd = None

REPO_DIR = Path(__file__).resolve().parent.parent
PREFIX = "factsheet_"

# 행 이름 → 지표. 앞뒤 공백·글머리 기호를 눌러서 견준다.
LABELS = {
    "total wireless service subscribers": "subs_wireless",
    "total wireless service": "subs_wireless",
    "mno": "subs_mno",
    "mvno": "subs_mvno",
    "5g handset": "subs_5g",
    "arpu": "arpu_wireless",
    "telephony (pstn+voip)": "subs_telephony",
    "pstn": "subs_pstn",
    "voip": "subs_voip",
    "broadband": "subs_broadband",
    "genie tv": "subs_iptv",
    "iptv": "subs_iptv",
    "olleh tv": "subs_iptv",
}

# 지표가 들어갈 만한 범위(명, ARPU 는 원). 단위 판정과 헛값 거르기에 쓴다.
RANGE = {
    "subs_wireless":  (12_000_000, 40_000_000),
    "subs_mno":       (10_000_000, 25_000_000),
    "subs_mvno":       (1_000_000, 15_000_000),
    "subs_5g":         (1_000_000, 20_000_000),
    "subs_telephony":  (5_000_000, 25_000_000),
    "subs_pstn":       (3_000_000, 25_000_000),
    "subs_voip":       (1_000_000, 10_000_000),
    "subs_broadband":  (3_000_000, 12_000_000),
    "subs_iptv":       (1_000_000, 12_000_000),
    "arpu_wireless":      (25_000, 60_000),      # 원
}
SCALES = (1.0, 1000.0)      # 명 그대로 / 천명으로 적힌 것

_Q = re.compile(r"^([1-4])\s*Q\s*'?(\d{2}|\d{4})$", re.I)
EXCEL_EPOCH = dt.datetime(1899, 12, 30)


def norm(s) -> str:
    t = re.sub(r"[\s ]+", " ", str(s or "")).strip().lower()
    t = t.lstrip("o•*※-–— ").strip()
    t = re.sub(r"\s*\(krw\)$|\s*\*+$", "", t).strip()
    return t


def parse_header(cell) -> tuple[str, str]:
    """머리글 한 칸 → `(period, period_type)`. 분기 문자열·날짜·일련번호를 다 받는다."""
    if isinstance(cell, dt.datetime):
        return f"{cell.year}-{cell.month:02d}", "month"
    if isinstance(cell, dt.date):
        return f"{cell.year}-{cell.month:02d}", "month"
    if isinstance(cell, (int, float)) and 20000 < cell < 60000:
        # 엑셀 일련번호. 1900년 윤년 버그 때문에 기준일이 1899-12-30 이다.
        #
        # **정수일 때만 받는다.** 안 그러면 ARPU 값이 날짜로 읽힌다 —
        # 31,047원이 이 범위 한가운데다. 실제로 그것 때문에 ARPU 줄 자체가
        # 머리글로 오인돼 ARPU 가 통째로 안 잡혔다. 날짜 머리글은 달의 1일이라
        # 소수점이 없다.
        if abs(cell - round(cell)) > 1e-6:
            return "", ""
        d = EXCEL_EPOCH + dt.timedelta(days=float(round(cell)))
        return f"{d.year}-{d.month:02d}", "month"
    t = re.sub(r"\s+", " ", str(cell or "")).strip()
    m = _Q.match(t)
    if m:
        y = m.group(2)
        year = int(y) if len(y) == 4 else 2000 + int(y)
        return f"{year}Q{m.group(1)}", "quarter"
    return "", ""


def detect_scale(value, metric: str) -> float | None:
    """값 × 배수 = 명(또는 원). 정할 수 없으면 None — 그 값은 버린다."""
    if not isinstance(value, (int, float)) or value <= 0:
        return None
    lo, hi = RANGE[metric]
    if metric == "arpu_wireless":
        return 1.0 if lo <= value <= hi else None
    hits = [s for s in SCALES if lo <= value * s <= hi]
    return hits[0] if len(hits) == 1 else None


def read_rows(path: Path) -> dict[str, list[list]]:
    """시트 이름 → 행 목록. `.xlsx` 와 옛 `.xls` 를 함께 다룬다."""
    p = str(path)
    out: dict[str, list[list]] = {}
    if p.lower().endswith((".xlsx", ".xlsm")):
        wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
        for name in wb.sheetnames:
            out[name] = [list(r) for r in wb[name].iter_rows(values_only=True)]
        wb.close()
    elif p.lower().endswith(".xls"):
        if xlrd is None:
            raise RuntimeError("xlrd 가 없다:  .venv/bin/pip install xlrd")
        wb = xlrd.open_workbook(p)
        for sh in wb.sheets():
            out[sh.name] = [[sh.cell_value(i, j) for j in range(sh.ncols)]
                            for i in range(sh.nrows)]
    else:
        raise RuntimeError(f"다룰 수 없는 형식: {path.suffix}")
    return out


def read_sheet(rows: list[list], report_id: str, warn: list) -> list[dict]:
    """시트 하나에서 수치를 뽑는다. 머리글 줄을 찾고 그 아래를 읽는다."""
    facts: list[dict] = []
    headers = []
    for i, r in enumerate(rows):
        cols = [(j, *parse_header(c)) for j, c in enumerate(r[1:], start=1)]
        cols = [(j, p, pt) for j, p, pt in cols if p]
        # **기간이 왼쪽에서 오른쪽으로 늘어나야 머리글이다.** 숫자 줄이 우연히
        # 날짜처럼 읽히는 일이 있는데(정수 ARPU 등), 그런 값은 오르내려서 여기서
        # 걸린다. 머리글 날짜는 예외 없이 오름차순이다.
        if len(cols) >= 3 and all(a[1] < b[1] for a, b in zip(cols, cols[1:])):
            headers.append((i, cols))
    if not headers:
        return []

    for idx, (hi, cols) in enumerate(headers):
        end = headers[idx + 1][0] if idx + 1 < len(headers) else len(rows)
        for r in rows[hi + 1:end]:
            if not r:
                continue
            metric = LABELS.get(norm(r[0]))
            if metric is None:
                continue
            for j, period, ptype in cols:
                v = r[j] if j < len(r) else None
                if not isinstance(v, (int, float)) or v <= 0:
                    continue
                scale = detect_scale(v, metric)
                if scale is None:
                    warn.append(f"{report_id}: {metric} {period} 값이 범위 밖({v:,.0f}) — 버림")
                    continue
                facts.append({
                    "report_id": report_id, "period": period, "period_type": ptype,
                    "basis": "시점", "consolidation": "별도", "accounting": "",
                    "metric_id": metric, "label_raw": str(r[0]).strip(),
                    "value": round(v * scale, 2),
                    "unit": "원" if metric == "arpu_wireless" else "명",
                    "page": "", "confidence": "auto",
                })
    return facts


def dedupe(facts: list[dict], reports: list[dict]) -> list[dict]:
    """같은 값이 여러 보고서에 반복되는 것을 접는다.

    **Factsheet 는 파일마다 과거 전체가 실려 있다.** 2011년 이후 177개 파일이
    저마다 2005년치까지 들고 있어서, 같은 달 초고속인터넷 가입자 수가 그대로
    177번 들어온다. 접기 전 88,505개 중 대부분이 이런 복사본이다.

    **값이 달라질 때만 남긴다** — 값이 바뀌었다는 것은 KT 가 수치를 고쳤다는
    뜻이고(파일에도 `Retroactively adjusted` 라고 적혀 있다), 그건 재작성이라
    보존할 값어치가 있다. 같은 값이면 **가장 최근 보고서** 것 하나만 둔다.
    """
    order = {r["report_id"]: r.get("published_at") or "" for r in reports}
    best: dict[tuple, dict] = {}
    for f in facts:
        key = (f["metric_id"], f["period"], f["consolidation"], round(f["value"], 2))
        cur = best.get(key)
        if cur is None or order.get(f["report_id"], "") > order.get(cur["report_id"], ""):
            best[key] = f
    return list(best.values())


def main() -> None:
    ap = argparse.ArgumentParser(description="IR Factsheet → facts.csv")
    ap.add_argument("--report", action="store_true", help="파일을 쓰지 않고 진단만")
    ap.add_argument("--limit", type=int)
    args = ap.parse_args()

    import csv
    reports = [r for r in csv.DictReader((REPO_DIR / "data" / "reports.csv").open(encoding="utf-8"))
               if r["doc_kind"] == "factsheet" and r["local_path"]]
    reports.sort(key=lambda r: r["period"])
    if args.limit:
        reports = reports[-args.limit:]

    facts: list[dict] = []
    warn: list[str] = []
    done = skipped_old = failed = 0

    for r in reports:
        path = REPO_DIR / r["local_path"]
        try:
            sheets = read_rows(path)
        except Exception as e:
            warn.append(f"{r['report_id']}: 열기 실패 {type(e).__name__}")
            failed += 1
            continue
        names = {n.strip().lower(): n for n in sheets}
        # 현대 형식만 다룬다. 나머지는 세어서 알린다.
        target = [names[k] for k in ("wireless(q)", "fixed(m)") if k in names]
        if not target:
            skipped_old += 1
            continue
        got = []
        for name in target:
            got.extend(read_sheet(sheets[name], r["report_id"], warn))
        facts.extend(got)
        done += 1

    print(f"Factsheet {len(reports)}건 — 읽음 {done} · 옛 형식 건너뜀 {skipped_old} · 실패 {failed}")
    print(f"수치 {len(facts):,}개")

    if warn:
        print(f"\n경고 {len(warn)}건 (앞 10)")
        for w in warn[:10]:
            print("   ", w)

    if facts:
        import collections
        by = collections.Counter(f["metric_id"] for f in facts)
        periods = sorted({f["period"] for f in facts})
        print(f"\n기간 {periods[0]} ~ {periods[-1]} ({len(periods)}개)")
        print("지표별:", ", ".join(f"{k} {v:,}" for k, v in by.most_common()))

    facts = dedupe(facts, reports)
    print(f"중복을 접은 뒤: {len(facts):,}개")

    if args.report:
        print("\n(--report 라 파일을 쓰지 않았다)")
        return

    removed, added = FS.replace_facts(PREFIX, facts)
    print(f"\nfacts.csv — factsheet 줄 {removed:,}개 걷어내고 {added:,}개 넣음")


if __name__ == "__main__":
    main()
