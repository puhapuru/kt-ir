#!/usr/bin/env python3
"""Financial Indicator(XLSX)에서 수치를 뽑아 `data/facts.csv` 를 만든다.

    .venv/bin/python scripts/extract_findicator.py            # 뽑아서 쓴다
    .venv/bin/python scripts/extract_findicator.py --report   # 안 쓰고 진단만

IR 자료 중 **여기부터 하는 것이 맞다.** 표라서 PDF 처럼 좌표를 더듬을 필요가
없고, 2011Q4~2026Q2 59개 분기가 이미 다 받아져 있다.

## 시트 모양 (2011~2026 내내 같다)

    2행  KT Financial Indicators (Separate)
    3행  (Billion KRW)                     ← **믿으면 안 된다. 아래 참고**
    4행                  K-GAAP | K-GAAP | K-IFRS      ← 회계기준
    5행  Financial Highlights | 2008 | 2009 | 2010 | 2010   ← 기간 머리글
    6행  Revenue              | 18932.8 | ...

시트는 둘이다 — `Separate`(옛 이름 `KT Only`) = 별도, `Consolidated` = 연결.

## 함정 셋 (전부 실제로 밟았다)

**1. 단위가 시트마다·연도마다 다르다.** 3행은 언제나 `(Billion KRW)` 라고
적혀 있지만 **거짓이다.** Consolidated 시트는 2016~2021 어느 시점에 백만원으로
바뀐다:

    2016Q4 Consolidated Revenue     23,790.3   ← 십억원
    2021Q4 Consolidated Revenue  5,710,200.0   ← 백만원 (1000배)

머리글을 믿고 고정했으면 절반이 1000배 틀렸을 것이다. 그래서 **Revenue 값의
크기로 판정한다.** KT 매출은 연 15~28조·분기 4~7조라, 십억원이면 4천~2만8천,
백만원이면 400만~2800만이다. 1000배 차이라 경계가 넉넉하다.

**2. 같은 연도가 두 번 나온다.** 2011Q4 파일의 2010년은 `K-GAAP 20233.5` 와
`K-IFRS 19918.4` 로 **둘 다 실려 있다.** 회계기준이 바뀌며 재작성된 것이다.
4행의 기준을 `accounting` 칸에 함께 적어야 구별된다. 안 적으면 같은 분기에
값이 두 개인 이유를 나중에 알 수 없다.

**3. 파일 하나에 여러 분기가 들어 있다.** 2026Q2 자료에 2022년부터 실려 있다.
그래서 같은 분기가 여러 보고서에 나오는데, **그것이 바로 재작성 이력**이다.
덮어쓰지 말고 `report_id` 를 달아 다 남긴다.

## 단위는 억원으로 맞춘다

`metrics.csv` 의 표준 단위가 억원이다. 십억원×10, 백만원÷100 — 둘 다 정확히
나누어떨어져 값이 상하지 않는다.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import factstore as FS

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl 이 없다:  python3 -m venv .venv && .venv/bin/pip install openpyxl")

REPO_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_DIR / "data"
REPORTS_CSV = DATA_DIR / "reports.csv"
FACTS_CSV = DATA_DIR / "facts.csv"

FACTS_COLUMNS = [
    "report_id", "period", "period_type", "basis", "consolidation", "accounting",
    "metric_id", "label_raw", "value", "unit", "page", "confidence",
]

# 시트 이름 → 연결/별도. 옛 파일은 `KT Only` 라고 쓴다.
SHEETS = {"separate": "별도", "kt only": "별도", "consolidated": "연결"}

# 뽑을 지표. **여기 없는 표기는 버리고 `--report` 로 알린다** — 뜻을 모르는
# 줄을 넘겨짚어 넣는 것이 빈칸보다 나쁘다.
LABELS = {
    "revenue": "revenue_total",
    "- service revenue": "revenue_service",
    "- service": "revenue_service",
    "- handset revenue": "revenue_goods",
    "- merchandise": "revenue_goods",
    "- sale of goods": "revenue_goods",
    "operating income": "operating_profit",
    "net income": "net_profit",
    "net income attributable to kt": "net_profit_controlling",
    "ebitda": "ebitda",
    "capex": "capex_total",
    "- access network": "capex_access",
    "- backbone network": "capex_backbone",
    "- b2b": "capex_b2b",
    "depreciation": "opex_depreciation",
}

# 금액이 아닌 것(비율·주식수)은 이번 판에서 뽑지 않는다. 단위 판정이 따로
# 필요해서, 섞으면 검산이 무의미해진다.
SKIP = re.compile(r"margin|ratio|rate|per share|outstanding|treasury|price|payout|"
                  r"eps|per\b|ev/|roe|roa|equity|to ebitda", re.I)

_Q = re.compile(r"^([1-4])\s*Q\s*'?(\d{2}|\d{4})$", re.I)
_Y = re.compile(r"^(19|20)\d{2}$")

# 지표마다 **연간 값이 들어갈 만한 범위(억원)**. 단위를 칸마다 정하는 잣대다.
#
# 왜 이런 표가 필요한가 — 단위가 **칸마다** 다르기 때문이다. 2020Q4 별도 시트
# 실측:
#
#     Revenue       2015 = 16942.4      2016 = 17,028,868   ← 같은 줄인데 바뀐다
#     Capex         2015 = 2397         2016 = 2359         ← 이 줄은 안 바뀐다
#     Depreciation  2015 = 3010.2       2016 = 3,005,383    ← 이 줄은 바뀐다
#
# 열로도, 행으로도 못 가른다. 그래서 **값 하나하나**를 보고 "십억원으로 읽어야
# 말이 되는가, 백만원으로 읽어야 말이 되는가" 를 판단한다. 두 후보가 1000배
# 차이라 범위를 넉넉히 잡아도 겹치지 않는다.
#
# 절댓값으로 본다 — 영업이익·순이익은 적자 해가 있다(2014년 별도 -7,194억).
METRIC_RANGE_억 = {
    "revenue_total":          (120_000, 320_000),
    "revenue_service":        (100_000, 260_000),
    "revenue_goods":           (10_000,  70_000),
    "operating_profit":         (2_000,  35_000),
    "net_profit":                 (500,  30_000),
    "net_profit_controlling":     (500,  30_000),
    "ebitda":                  (18_000,  65_000),
    "capex_total":             (12_000,  50_000),
    "capex_access":             (6_000,  32_000),
    "capex_backbone":           (2_000,  10_000),
    "capex_b2b":                (1_500,  10_000),
    "opex_depreciation":       (20_000,  45_000),
}
# 분기 값은 연간의 1/4 언저리지만 계절성이 크다(4분기에 CAPEX 가 몰린다).
QUARTER_LO, QUARTER_HI = 1 / 8, 3 / 4


def norm(s) -> str:
    """표기 흔들림을 눌러 준다. `- Service revenue` → `- service revenue`."""
    t = re.sub(r"\s+", " ", str(s or "")).strip().lower()
    t = t.replace("–", "-").replace("*", "").strip()
    return t


def parse_period(cell) -> tuple[str, str]:
    """머리글 한 칸에서 `(period, period_type)`. 못 읽으면 `("", "")`."""
    t = re.sub(r"\s+", " ", str(cell or "")).strip()
    m = _Q.match(t)
    if m:
        q, y = m.group(1), m.group(2)
        year = int(y) if len(y) == 4 else 2000 + int(y)
        return f"{year}Q{q}", "quarter"
    if _Y.match(t):
        return t, "year"
    return "", ""


def detect_scale(value, period_type: str, metric_id: str) -> float | None:
    """`값 × 배수 = 억원` 이 되는 배수. 정할 수 없으면 None.

    **칸 하나마다 부른다.** 열이나 행 단위로 정하면 안 된다 — 같은 줄 안에서도
    연도에 따라 단위가 바뀌는 자료가 있다(위 `METRIC_RANGE_억` 설명 참고).

    후보 둘(십억원·백만원)이 1000배 차이라, 그 지표가 들어갈 만한 범위에
    떨어지는 쪽은 보통 하나뿐이다. **둘 다 맞거나 둘 다 안 맞으면 None 을
    돌려 그 값을 버린다** — 넘겨짚느니 빠지는 편이 낫다.
    """
    if not isinstance(value, (int, float)) or value == 0:
        return None
    rng = METRIC_RANGE_억.get(metric_id)
    if rng is None:
        return None
    lo, hi = rng
    if period_type == "quarter":
        lo, hi = lo * QUARTER_LO, hi * QUARTER_HI

    # 후보는 **둘뿐**이다 — 십억원(×10)과 백만원(÷100). 1000배 차이라 범위가
    # 넉넉해도 겹치지 않는다.
    #
    # 억원(×1)을 후보에 넣어 봤다가 되돌렸다. 십억원과 10배밖에 차이가 안 나서
    # 순이익처럼 범위가 넓은 지표에서 **두 후보가 동시에 걸려** 판정 불가가
    # 무더기로 생기고, 그 자리를 이웃 칸으로 메우다 1000배 오류가 되살아났다
    # (단위의심 0건 → 10건). 후보를 늘리는 것이 개선이라는 보장이 없다.
    hits = [m for m in (10.0, 0.01) if lo <= abs(value) * m <= hi]
    return hits[0] if len(hits) == 1 else None


def read_sheet(ws, consolidation: str, report_id: str, warn: list) -> list[dict]:
    """시트 하나에서 수치를 뽑는다."""
    rows = list(ws.iter_rows(values_only=True))
    facts: list[dict] = []

    # 머리글 줄을 찾는다 — B칸이 섹션 이름이고 C칸부터 기간이 적힌 줄.
    header_rows = []
    for i, r in enumerate(rows):
        if len(r) < 4:
            continue
        periods = [parse_period(c) for c in r[2:]]
        if sum(1 for p, _ in periods if p) >= 2:
            header_rows.append(i)
    if not header_rows:
        warn.append(f"{report_id}/{consolidation}: 기간 머리글을 못 찾음")
        return []

    for hi in header_rows:
        header = rows[hi]
        cols = [(j + 2, *parse_period(c)) for j, c in enumerate(header[2:])]
        cols = [(j, p, pt) for j, p, pt in cols if p]

        # 바로 윗줄이 회계기준(K-IFRS/K-GAAP)이면 칸마다 받아 둔다.
        acct = {}
        if hi > 0:
            above = rows[hi - 1]
            for j, _, _ in cols:
                v = above[j] if j < len(above) else None
                if v and re.search(r"k-?(ifrs|gaap)", str(v), re.I):
                    acct[j] = re.sub(r"\s+", " ", str(v)).strip()

        # 섹션 본문 — 다음 머리글 전까지
        end = next((h for h in header_rows if h > hi), len(rows))
        for r in rows[hi + 1:end]:
            if len(r) < 3:
                continue
            label = norm(r[1])
            if not label or SKIP.search(label):
                continue
            metric = LABELS.get(label)
            if metric is None:
                if any(isinstance(r[j], (int, float)) for j, _, _ in cols):
                    warn.append(f"{report_id}/{consolidation}: 모르는 표기 {r[1]!r}")
                continue
            # 단위는 **칸마다** 정한다. 열이나 행으로 묶으면 틀린다.
            cell_scale = {}
            for j, period, ptype in cols:
                v = r[j] if j < len(r) else None
                if isinstance(v, (int, float)) and v != 0:
                    cell_scale[j] = detect_scale(v, ptype, metric)

            # **못 정한 칸은 버린다.** 같은 줄의 이웃 칸 단위를 빌려 메우는
            # 방식을 넣어 봤다가 걷어냈다 — 빈칸은 없어지지만 **틀린 값**이
            # 들어간다(2015Q1 연결 영업이익 분기합이 234,232 로 부풀었다.
            # 연간은 -3,665 다). 빠진 값은 검산에 안 걸리지만 틀린 값은
            # 분석을 통째로 망친다. 버린 것은 아래 경고로 남으니 사람이
            # 확인해 채우면 된다.
            for j, period, ptype in cols:
                v = r[j] if j < len(r) else None
                if not isinstance(v, (int, float)) or v == 0:
                    continue
                scale = cell_scale.get(j)
                if scale is None:
                    warn.append(
                        f"{report_id}/{consolidation}: {metric} {period} "
                        f"단위 미정({v:,.1f}) — 버림")
                    continue
                facts.append({
                    "report_id": report_id,
                    "period": period,
                    "period_type": ptype,
                    # 분기 칸은 단일분기다(4개 합이 연간과 맞는 것을 검산한다).
                    "basis": "단일분기" if ptype == "quarter" else "연간",
                    "consolidation": consolidation,
                    "accounting": acct.get(j, ""),
                    "metric_id": metric,
                    "label_raw": str(r[1]).strip(),
                    "value": round(v * scale, 2),
                    "unit": "억원",
                    "page": "",
                    "confidence": "auto",
                })
    return facts


def check_sums(facts: list[dict]) -> list[str]:
    """분기 4개 합이 연간과 맞는지 본다. **단위·칸 오독을 잡는 그물이다.**"""
    bucket = defaultdict(dict)
    for f in facts:
        key = (f["report_id"], f["consolidation"], f["accounting"], f["metric_id"])
        bucket[key][f["period"]] = f["value"]

    problems = []
    for (rid, cons, acct, metric), by_period in bucket.items():
        for period, annual in list(by_period.items()):
            if "Q" in period:
                continue
            qs = [by_period.get(f"{period}Q{q}") for q in (1, 2, 3, 4)]
            if any(v is None for v in qs) or not annual:
                continue
            total = sum(qs)
            if abs(total - annual) <= max(abs(annual) * 0.01, 1.0):
                continue
            # **1000배급 어긋남과 몇 % 어긋남은 성격이 전혀 다르다.**
            # 앞은 단위 오독(추출 잘못)이고, 뒤는 KT 가 연간만 재작성해
            # 분기 내역과 안 맞는 것이다(자료 자체가 그렇다).
            r = abs(total / annual) if annual else 0
            kind = "단위의심" if (r > 100 or (r and r < 0.01)) else "값차이"
            problems.append(
                f"[{kind}] {rid}/{cons}/{metric} {period}: "
                f"분기합 {total:,.0f} ≠ 연간 {annual:,.0f}")
    return problems


def main() -> None:
    ap = argparse.ArgumentParser(description="Financial Indicator XLSX → facts.csv")
    ap.add_argument("--report", action="store_true", help="파일을 쓰지 않고 진단만 한다")
    ap.add_argument("--limit", type=int, help="앞에서 N개 보고서만")
    args = ap.parse_args()

    reports = [r for r in csv.DictReader(REPORTS_CSV.open(encoding="utf-8"))
               if r["doc_kind"] == "financial_indicator" and r["local_path"]]
    reports.sort(key=lambda r: r["period"])
    if args.limit:
        reports = reports[:args.limit]

    facts: list[dict] = []
    warn: list[str] = []
    failed = 0
    for r in reports:
        path = REPO_DIR / r["local_path"]
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception as e:
            warn.append(f"{r['report_id']}: 열기 실패 {type(e).__name__}")
            failed += 1
            continue
        for name in wb.sheetnames:
            cons = SHEETS.get(norm(name))
            if cons is None:
                continue
            facts.extend(read_sheet(wb[name], cons, r["report_id"], warn))
        wb.close()

    print(f"보고서 {len(reports)}건 · 열기 실패 {failed}건 · 수치 {len(facts):,}개")

    unknown = sorted({w.split("모르는 표기 ")[1] for w in warn if "모르는 표기" in w})
    if unknown:
        print(f"\n지도에 없는 표기 {len(unknown)}종 (버렸다):")
        for u in unknown[:20]:
            print("   ", u)

    other = [w for w in warn if "모르는 표기" not in w]
    if other:
        print(f"\n그 밖의 경고 {len(other)}건:")
        for w in other[:12]:
            print("   ", w)

    problems = check_sums(facts)
    unit_bad = [p for p in problems if p.startswith("[단위의심]")]
    val_diff = [p for p in problems if p.startswith("[값차이]")]
    print(f"\n검산 — 분기합 vs 연간")
    print(f"  단위의심(추출 잘못): {len(unit_bad)}건 {'✓' if not unit_bad else '← 고쳐야 한다'}")
    print(f"  값차이(자료 자체)  : {len(val_diff)}건")
    for p in unit_bad[:8]:
        print("   ", p)
    for p in val_diff[:5]:
        print("   ", p)

    if args.report:
        print("\n(--report 라 파일을 쓰지 않았다)")
        return

    # **자기 몫만 갈아 끼운다.** 파일을 통째로 덮어쓰면 DART 등 다른 추출기가
    # 넣어 둔 줄이 사라진다.
    removed, added = FS.replace_facts("financial_indicator_", facts)
    print(f"\nfacts.csv — 기존 {removed:,}줄 걷어내고 {added:,}줄 넣음")


if __name__ == "__main__":
    main()
